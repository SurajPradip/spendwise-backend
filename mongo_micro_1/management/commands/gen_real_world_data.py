from django.core.management.base import BaseCommand
from openpyxl import load_workbook
import os
from dateutil import parser
from datetime import datetime
from mongo_micro_1.models import SpendwiseBasicDetails

class Command(BaseCommand):
    help = "Load Real World Data"

    def find_num_rows(self, num_columns, worksheet):
        def check_if_emp(row, num_columns):
            row_list_1 = []
            for column in range(1, num_columns + 1):
                row_list_1.append(worksheet.cell(row=row, column=column))
            if all(item.value is None for item in row_list_1):
                return True
            else:
                return False

        row_no = 1
        empty_no = 0
        while empty_no < 10:
            if not check_if_emp(row_no, num_columns):
                row_no += 1
                empty_no = 0
            else:
                empty_no += 1
                row_no += 1
        return row_no - empty_no - 1
    
    category_map = {
        "Food":1,
        "Apartment":2,
        "Petrol":3,
        "Regret":4,
        "Luxury":5,
        "Others":6,
        "Investment":7,
        "Home":8,
        "Debt":9,
        "Travel":10,
        "Self Care":11,
        "Donations":12
        
    }

    def handle(self, *args, **options):
        module_dir = os.path.dirname(__file__)
        file_path = os.path.join(module_dir, 'excel_file.xlsx')
        
        wb_obj = load_workbook(file_path)
        sheet = wb_obj.worksheets[5]

        row_count = self.find_num_rows(4, sheet)
        for row in sheet.iter_rows(min_row=2, max_row=row_count, values_only=True):
            date_string = row[0]
            reason = row[1]
            category = row[2]
            price = row[3]
            s_faction = row[4]

            try:
                date_obj = parser.parse(str(date_string))
                formatted_date = date_obj.strftime('%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                formatted_date = "Invalid Date"
            if category:
                SpendwiseBasicDetails.objects.create(date=formatted_date,reason=reason,category=self.category_map[category],price=price,s_faction=s_faction)
                print("created...")

